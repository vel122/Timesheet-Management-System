from datetime import date

import frappe
import pandas as pd
from frappe.utils import add_days, formatdate
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


@frappe.whitelist()
def generate_missing_timesheet_excel():
	"""Generate formatted missing timesheet Excel with Draft column"""
	today = date.today()
	week_start = add_days(today, -7)
	week_end = today

	# employees = frappe.get_all(
	#     "Employee",
	#     filters={"status": "Active"},
	#     fields=["name","employee_name"],
	# )
	employees = frappe.db.sql("SELECT name, employee_name FROM tabEmployee", as_dict=True)
	print("employees", employees)

	summary = []

	for emp in employees:
		missing_dates = []
		draft_dates = []

		for i in range(7):
			current_date = add_days(week_start, i)

			# Check if timesheet exists
			timesheet = frappe.db.get_all(
				"Timesheet",
				filters={
					"employee": emp.name,
					"start_date": ["<=", current_date],
					"end_date": [">=", current_date],
				},
				fields=["name", "status"],
			)

			if not timesheet:
				# ❌ No timesheet found
				missing_dates.append(formatdate(current_date))
			else:
				# 📝 If draft
				if any(t.status == "Draft" for t in timesheet):
					draft_dates.append(formatdate(current_date))
				# ✅ Submitted ones ignored

		if missing_dates or draft_dates:
			summary.append(
				{
					"Employee ID": emp.name,
					"Employee Name": emp.employee_name,
					"Missing Dates": ", ".join(missing_dates) if missing_dates else "",
					"Draft Dates": ", ".join(draft_dates) if draft_dates else "",
				}
			)

	if not summary:
		frappe.msgprint("No missing or draft timesheets found for this week.")
		return None

	# Convert to DataFrame
	df = pd.DataFrame(summary)
	file_name = f"Missing_Timesheet_{formatdate(week_start)}_to_{formatdate(week_end)}.xlsx"
	file_path = frappe.utils.get_site_path("public", "files", file_name)

	# Write Excel (without index)
	df.to_excel(file_path, index=False)

	# ----------------- Apply Formatting -----------------
	wb = load_workbook(file_path)
	ws = wb.active

	header_fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
	header_font = Font(bold=True, color="FFFFFF")
	border_style = Border(
		left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
	)

	# Header styling
	for cell in ws[1]:
		cell.fill = header_fill
		cell.font = header_font
		cell.alignment = Alignment(horizontal="center", vertical="center")
		cell.border = border_style

	# Data styling
	for row in ws.iter_rows(min_row=2):
		for cell in row:
			cell.border = border_style
			cell.alignment = Alignment(vertical="center", wrap_text=True)

	# Auto-adjust column widths
	for col in ws.columns:
		max_length = 0
		column = col[0].column_letter
		for cell in col:
			if cell.value:
				max_length = max(max_length, len(str(cell.value)))
		ws.column_dimensions[column].width = max_length + 3

	ws.freeze_panes = "A2"
	wb.save(file_path)

	file_doc = frappe.get_doc(
		{"doctype": "File", "file_name": file_name, "is_private": 0, "file_url": f"/files/{file_name}"}
	)
	file_doc.insert(ignore_permissions=True)
	# frappe.db.commit()
	return file_doc.file_url


def get_employee_list():
	employees = frappe.get_all(
		"Employee",
		filters={"status": "Active"},
		fields=["name", "employee_name"],
	)
	return employees
